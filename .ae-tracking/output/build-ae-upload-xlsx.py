import json
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / ".ae-tracking" / "output"
DRAFT = OUT_DIR / "shuffla-tracking-plan-draft.json"
TEMPLATE = Path("/opt/homebrew/lib/node_modules/ae-tracking/tracking-plan-template/TE官方模板_dataTrackSample.xlsx")
OUT = OUT_DIR / (sys.argv[1] if len(sys.argv) > 1 else "shuffla-tracking-plan-ae-upload.xlsx")

TYPE_TO_CN = {
    "string": "文本",
    "number": "数值",
    "bool": "布尔",
    "datetime": "时间",
    "array_row": "对象组",
    "array_string": "列表",
}


def ascii_text(value):
    if value is None:
        return ""
    text = str(value)
    encoded = text.encode("ascii", "ignore").decode("ascii").strip()
    return encoded or text


def title_from_name(name):
    return " ".join(part.upper() if part in {"id", "url", "ai", "sdk"} else part.capitalize() for part in name.split("_"))


def reset_sheet(ws):
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def write_rows(ws, rows):
    reset_sheet(ws)
    for row in rows:
        ws.append(row)


def main():
    draft = json.loads(DRAFT.read_text(encoding="utf-8"))
    shutil.copyfile(TEMPLATE, OUT)
    wb = load_workbook(OUT)

    prop_by_name = {prop["name"]: prop for prop in draft["event_properties"]}
    ascii_mode = OUT.name.endswith("-ascii.xlsx")

    event_ws = wb["#事件数据"]
    reset_sheet(event_ws)
    event_ws.append([
        "事件名（必填）",
        "事件显示名",
        "事件说明",
        "事件标签",
        "属性名（必填）",
        "属性显示名",
        "属性类型（必填）",
        "属性说明",
    ])
    for event in draft["events"]:
        start_row = event_ws.max_row + 1
        prop_names = event.get("prop_names") or [None]
        for prop_name in prop_names:
            prop = prop_by_name.get(prop_name) if prop_name else None
            event_ws.append([
                event["event_name"],
                title_from_name(event["event_name"]) if ascii_mode else event.get("display_name", ""),
                title_from_name(event["event_name"]) if ascii_mode else event.get("event_desc", ""),
                event.get("event_tag", ""),
                prop.get("name", "") if prop else "",
                title_from_name(prop["name"].replace(".", "_")) if prop and ascii_mode else (prop.get("display_name", "") if prop else ""),
                TYPE_TO_CN.get(prop.get("type", ""), "") if prop else "",
                title_from_name(prop["name"].replace(".", "_")) if prop and ascii_mode else (prop.get("desc", "") if prop else ""),
            ])
        end_row = event_ws.max_row
        if end_row > start_row:
            for col in ("A", "B", "C", "D"):
                event_ws.merge_cells(f"{col}{start_row}:{col}{end_row}")

    write_rows(
        wb["#公共事件属性"],
        [["属性名（必填）", "属性显示名", "属性类型（必填）", "属性说明"]]
        + [
            [
                prop["name"],
                title_from_name(prop["name"]) if ascii_mode else prop.get("display_name", ""),
                TYPE_TO_CN[prop["type"]],
                title_from_name(prop["name"]) if ascii_mode else prop.get("desc", ""),
            ]
            for prop in draft["common_event_properties"]
        ],
    )

    write_rows(
        wb["#用户数据"],
        [["属性名（必填）", "属性显示名", "属性类型（必填）", "更新方式", "属性说明", "属性标签"]]
        + [
            [
                prop["name"],
                title_from_name(prop["name"]) if ascii_mode else prop.get("display_name", ""),
                TYPE_TO_CN[prop["type"]],
                prop.get("update_type", "user_set"),
                title_from_name(prop["name"]) if ascii_mode else prop.get("desc", ""),
                prop.get("prop_tag", ""),
            ]
            for prop in draft["user_properties"]
        ],
    )

    id_ws = wb["#用户ID体系"]
    write_rows(
        id_ws,
        [
            ["游戏类型", "属性名", "属性显示名", "属性说明", "赋值说明", None, None, None, None, None, None],
            ["单账号单角色", "#account_id", "账户ID", "系统属性", "设置为 user.id", None, None, None, None, None, None],
            [None, "#distinct_id", "访客ID", "系统属性", "SDK 自动生成或设备相关 ID", None, None, None, None, None, None],
            ["单账号多角色", "#account_id", "账户ID", "系统属性", "设置为 user.id", None, None, None, None, None, None],
            [None, "#distinct_id", "访客ID", "系统属性", "SDK 自动生成或设备相关 ID", None, None, None, None, None, None],
        ],
    )
    for cell_range in ("A2:A3", "A4:A5", "E1:K1", "E2:K2", "E3:K3", "E4:K4", "E5:K5"):
        id_ws.merge_cells(cell_range)

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
