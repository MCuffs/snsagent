import copy
import json
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / ".ae-tracking" / "output"
DRAFT = OUT_DIR / "shuffla-tracking-plan-draft.json"
TEMPLATE = Path("/Users/jeongminsu/Downloads/Te 데이터 트래킹 샘플 템플릿 (1).xlsx")
OUT = OUT_DIR / (sys.argv[1] if len(sys.argv) > 1 else "shuffla-tracking-plan-ko-template.xlsx")

TYPE_TO_KO = {
    "string": "string",
    "number": "number",
    "bool": "boolean",
    "datetime": "time",
    "array_row": "object group",
    "array_string": "list",
}

UPDATE_TO_KO = {
    "user_set": "user_set",
    "user_setOnce": "user_setOnce",
    "user_add": "user_add",
}

AUTOTRACK_EVENTS = [
    {
        "event_name": "ta_pageview",
        "display_name": "페이지뷰 자동 수집",
        "event_desc": "ThinkingData JavaScript SDK autoTrack 호출 시 수집되는 페이지뷰 이벤트",
        "event_tag": "시스템",
        "prop_names": [],
    },
    {
        "event_name": "ta_page_show",
        "display_name": "페이지 표시 자동 수집",
        "event_desc": "ThinkingData JavaScript SDK pageShow 자동 수집 이벤트",
        "event_tag": "시스템",
        "prop_names": [],
    },
    {
        "event_name": "ta_page_hide",
        "display_name": "페이지 숨김 자동 수집",
        "event_desc": "ThinkingData JavaScript SDK pageHide 자동 수집 이벤트",
        "event_tag": "시스템",
        "prop_names": [],
    },
]


def clear_data_rows(ws, header_row=1):
    for merged in list(ws.merged_cells.ranges):
        min_row = merged.min_row
        if min_row > header_row:
            ws.unmerge_cells(str(merged))
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)


def copy_row_style(ws, source_row, target_row, max_col):
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.font:
            dst.font = copy.copy(src.font)
        if src.border:
            dst.border = copy.copy(src.border)


def append_styled(ws, values, style_row=2):
    ws.append(values)
    copy_row_style(ws, style_row, ws.max_row, len(values))


def main():
    draft = json.loads(DRAFT.read_text(encoding="utf-8"))
    shutil.copyfile(TEMPLATE, OUT)
    wb = load_workbook(OUT)

    prop_by_name = {prop["name"]: prop for prop in draft["event_properties"]}

    event_ws = wb["#이벤트 데이터"]
    clear_data_rows(event_ws)
    events = list(draft["events"]) + AUTOTRACK_EVENTS
    for event in events:
        start_row = event_ws.max_row + 1
        prop_names = event.get("prop_names") or [None]
        for prop_name in prop_names:
            prop = prop_by_name.get(prop_name) if prop_name else None
            append_styled(event_ws, [
                event["event_name"],
                event.get("display_name", ""),
                event.get("event_desc", ""),
                event.get("event_tag", ""),
                prop.get("name", "") if prop else None,
                prop.get("display_name", "") if prop else None,
                TYPE_TO_KO.get(prop.get("type", ""), "") if prop else None,
                prop.get("desc", "") if prop else None,
            ])
        end_row = event_ws.max_row
        if end_row > start_row:
            for col in ("A", "B", "C", "D"):
                event_ws.merge_cells(f"{col}{start_row}:{col}{end_row}")

    dedup_cross_pools = "dedup" in OUT.name
    event_prop_names = {prop["name"] for prop in draft["event_properties"]}
    common_prop_names = {prop["name"] for prop in draft["common_event_properties"]}

    common_ws = wb["#공통 이벤트 속성"]
    clear_data_rows(common_ws)
    for prop in draft["common_event_properties"]:
        if dedup_cross_pools and prop["name"] in event_prop_names:
            continue
        append_styled(common_ws, [
            prop["name"],
            prop.get("display_name", ""),
            TYPE_TO_KO[prop["type"]],
            prop.get("desc", ""),
        ])

    user_ws = wb["#유저 데이터"]
    clear_data_rows(user_ws)
    for prop in draft["user_properties"]:
        if dedup_cross_pools and (prop["name"] in event_prop_names or prop["name"] in common_prop_names):
            continue
        append_styled(user_ws, [
            prop["name"],
            prop.get("display_name", ""),
            TYPE_TO_KO[prop["type"]],
            UPDATE_TO_KO.get(prop.get("update_type", "user_set"), "user_set"),
            prop.get("desc", ""),
            prop.get("prop_tag", ""),
        ])

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
